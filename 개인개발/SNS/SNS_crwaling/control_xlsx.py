from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side, Color, PatternFill
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import DateAxis
import datetime


export_time = datetime.datetime.now()
day = export_time.date()



# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook('instagram.xlsx', data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['{}'.format(day)]



max_row = load_ws.max_row


load_ws.cell(1,10).value = "총 게시물"
load_ws.cell(1,11).value = "노출"
load_ws.cell(1,12).value = "도달"
load_ws.cell(1,13).value = "공감"
load_ws.cell(1,14).value = "댓글"


load_ws.cell(2,10).value = '={}'.format(max_row-1)
load_ws.cell(2,11).value = '=sum(c2:c{})'.format(max_row)
load_ws.cell(2,12).value = '=sum(d2:d{})'.format(max_row)
load_ws.cell(2,13).value = '=sum(e2:e{})'.format(max_row)
load_ws.cell(2,14).value = '=sum(f2:f{})'.format(max_row)


font_15 = Font(name='맑은 고딕', size=15, bold=True)
font_white = Font(color='ffffff',bold=True)
align_center = Alignment(horizontal='center', vertical='center')
align_vcenter = Alignment(vertical='center')

border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')

)

fill_orange = PatternFill(patternType='solid',fgColor=Color('ffe9b8'))
fill_blue = PatternFill(patternType='solid',fgColor=Color('000080'))

for row in load_ws['B1':'G1']:
    for cell in row:
        cell.alignment = align_center
        cell.fill = fill_orange
        cell.border = border_thin


for column in load_ws['A1':'A{}'.format(max_row)]:
    for cell in column:
        cell.alignment = align_center
        cell.fill = fill_orange
        cell.border = border_thin


for row in load_ws['J1':'N1']:
    for cell in row:
        cell.alignment = align_center
        cell.fill = fill_blue
        cell.border = border_thin
        cell.font = font_white

chart = BarChart()
chart.title = "게시물 인사이트 요약"
# chart.x_axis.title = "날짜"
# chart.y_axis.title = "변화량"
datas = Reference(load_ws, min_col=3,min_row=1,max_col=6,max_row=max_row)
chart.add_data(datas, from_rows=False, titles_from_data=True)

chart.width= 30
chart.height = 10

category = Reference(load_ws,min_col=7,min_row=2,max_col=7,max_row=max_row)
chart.set_categories(category)

load_ws.add_chart(chart, "J4")







#
# # 지정한 셀의 값 출력
#
# get_cells = load_ws['C2': 'C6']
# for row in get_cells:
#     for cell in row:
#         print(cell.value)
#
# # 모든 행 단위로 출력
#
# for row in load_ws.rows:
#     print(row)
#
# # 모든 열 단위로 출력
#
# for column in load_ws.columns:
#     print(column)
#
# # 모든 행과 열 출력
#
# all_values = []
# for row in load_ws.rows:
#     row_value = []
#     for cell in row:
#         row_value.append(cell.value)
#     all_values.append(row_value)
# print(all_values)


load_wb.save('C:/Users/USER/Desktop/instagram.xlsx')