import random
import time
import pandas as pd
from selenium import webdriver
import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side, Color, PatternFill
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference , BarChart
import sys




export_time = datetime.datetime.now()
day = export_time.date()



sns_name=[]
like_lst = []
comment_lst=[]
reach_lst =[]
exposure_lst=[]
date_lst = []



id = 'studiohollin@naver.com'
pw = 'ghffls2020!'


driver = webdriver.Chrome('chromedriver.exe')
driver.implicitly_wait(3)

# 인스타그램 로그인
driver.get('https://www.instagram.com/accounts/login/')

time.sleep(random.randint(2,4))

driver.find_element("name", 'username').send_keys(id)
driver.find_element("name", 'password').send_keys(pw)
time.sleep(random.randint(2,4))

driver.find_element("xpath", '//*[@id="loginForm"]/div/div[3]').click()


time.sleep(random.randint(2,5))


# 팝업 창 1 제거
driver.find_element("class name", "cmbtv").click()


time.sleep(random.randint(3,5))


# 팝업 창 2 제거
driver.find_element("class name", "_a9--._a9_1").click()

time.sleep(random.randint(3,5))

# 개인 프로필 진입
div = driver.find_element("class name","_aaav")
btn= div.find_element("tag name",'img')
btn.click()
driver.find_element("class name", "_ab8w._ab94._ab97._ab9f._ab9k._ab9p._ab9-._aba8._aaay._abcm").click()

time.sleep(10)





# 3열로 나눠짐 3개씩


indicator = driver.find_element("class name","_ac2a").text

indicator = int(indicator)-15

print("게시물 수:",indicator)

time.sleep(random.randint(2,5))

row_num = int(indicator)//3
row_num_remain = int(indicator)%3


if row_num_remain != 0:
    row_num += 1




time.sleep(random.randint(2,5))

# 게시물 클릭
post_1 = driver.find_element("class name",'_aabd._aa8k._aanf')
post_click = post_1.find_element("tag name", 'a').click()


time.sleep(random.randint(2,5))


def collertor(i,j):

    try:
        # 인사이트 클릭
        driver.find_element("xpath", '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/section[1]/div/section/button').click()
        time.sleep(random.randint(2,5))

        # 정보 수집
        like = driver.find_element("xpath",
                                   '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div/div/div/div[2]/div/div/div[2]/div/div/div[3]/div/div/div/div/div[1]/div[2]/span').text
        comment = driver.find_element("xpath",
                                      '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div/div/div/div[2]/div/div/div[2]/div/div/div[3]/div/div/div/div/div[2]/div[2]/span').text
        reach = driver.find_element("xpath",
                                    '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div/div/div/div[2]/div/div/div[2]/div/div/div[7]/div/div/div/div[2]/div/span[1]').text
        exposure = driver.find_element("xpath",
                                       '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div/div/div/div[2]/div/div/div[2]/div/div/div[7]/div/div/div/div[3]/span[2]').text
        date = driver.find_element("xpath",
                                   '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[2]/div/div/a/div/time').text


        # 정보 저장
        sns_name.append('인스타그램')
        like_lst.append(like)
        comment_lst.append(comment)
        reach_lst.append(reach)
        exposure_lst.append(exposure)
        date_lst.append(date)

        # 인사이트 정보 닫기
        # time.sleep(random.randint(2,5))
        driver.find_element("xpath",
                            '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div/div/div/div[1]/div/div[3]/div/button').click()


        # 다음게시물로 전환
        # time.sleep(random.randint(2, 5))
        if(i+1!=row_num or j!=row_num_remain):
            span_next = driver.find_element("class name", "_aaqg._aaqh")
            next_button = span_next.find_element("tag name", 'button')
            next_button.click()
        elif(i+1==row_num and j==row_num_remain):
            print("요약 완료")

            driver.close()


    except:
        # ddd = driver.find_element("xpath",
        #                           '/html/body/div[1]/div/div/div/div[2]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[1]/div/div/div[4]/button')
        #
        date = driver.find_element("xpath",
                                   '/html/body/div[1]/div/div[1]/div/div[2]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[2]/div/div/a/div/time').text

        sns_name.append('인스타그램')
        like_lst.append(0)
        comment_lst.append(0)
        reach_lst.append(0)
        exposure_lst.append(0)
        date_lst.append(date)

        span_next = driver.find_element("class name", "_aaqg._aaqh")
        next_button = span_next.find_element("tag name", 'button')
        next_button.click()


# 함수 실행
for i in range(0,row_num):
    if i+1 != row_num:
        for j in range(1,4):
            collertor(i,j)
    elif i+1==row_num:
        if row_num_remain == 0:
            row_num_remain = 3
        for j in range(1,row_num_remain+1):
            collertor(i,j)


### 데이터 프레임 csv파일 형태로 저장



instagram_record = pd.DataFrame({'활동영역': sns_name,
                                 '노출': exposure_lst,
                                    '도달': reach_lst,
                                    '공감': like_lst,
                                    '댓글': comment_lst,
                                 '게시일':date_lst})

print(instagram_record)

instagram_record = instagram_record.astype({'노출':'int','도달':'int','공감':'int','댓글':'int'})



### Save into worksheet




def xlsx_pre(ws):

    load_ws = ws

    max_row = load_ws.max_row

    load_ws.cell(1, 10).value = "총 게시물"
    load_ws.cell(1, 11).value = "노출"
    load_ws.cell(1, 12).value = "도달"
    load_ws.cell(1, 13).value = "공감"
    load_ws.cell(1, 14).value = "댓글"

    load_ws.cell(2, 10).value = '={}'.format(max_row - 1)
    load_ws.cell(2, 11).value = '=sum(c2:c{})'.format(max_row)
    load_ws.cell(2, 12).value = '=sum(d2:d{})'.format(max_row)
    load_ws.cell(2, 13).value = '=sum(e2:e{})'.format(max_row)
    load_ws.cell(2, 14).value = '=sum(f2:f{})'.format(max_row)




def xlsx_decoration(wb,ws):


    ws.cell(row=1,column=2).value = "활동영역"
    ws.cell(row=1,column=3).value = "노출"
    ws.cell(row=1,column=4).value = "도달"
    ws.cell(row=1,column=5).value = "공감"
    ws.cell(row=1,column=6).value = "댓글"
    ws.cell(row=1,column=7).value = "게시일"

    for i in range(indicator):
        ws.cell(row=i+2,column=1).value = i
        ws.cell(row=i+2,column=2).value = "인스타그램"
        ws.cell(row=i + 2, column=3).value = int(exposure_lst[i])
        ws.cell(row=i + 2, column=4).value = int(reach_lst[i])
        ws.cell(row=i + 2, column=5).value = int(like_lst[i])
        ws.cell(row=i + 2, column=6).value = int(comment_lst[i])
        ws.cell(row=i + 2, column=7).value = date_lst[i]




    load_wb = wb
    load_ws = ws

    max_row = load_ws.max_row


    load_ws.cell(1,10).value = "총 게시물"
    load_ws.cell(1,11).value = "노출"
    load_ws.cell(1,12).value = "도달"
    load_ws.cell(1,13).value = "공감"
    load_ws.cell(1,14).value = "댓글"


    load_ws.cell(2,10).value = '={}'.format(max_row-1)
    load_ws.cell(2,11).value = '=sum(c2:c{})'.format(max_row)
    load_ws.cell(2,12).value = '=sum(d2:d{})'.format(max_row)
    load_ws.cell(2,13).value = '=sum(e2:e{})'.format(max_row)
    load_ws.cell(2,14).value = '=sum(f2:f{})'.format(max_row)


    font_15 = Font(name='맑은 고딕', size=15, bold=True)
    font_white = Font(color='ffffff',bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    align_vcenter = Alignment(vertical='center')

    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')

    )

    fill_orange = PatternFill(patternType='solid',fgColor=Color('ffe9b8'))
    fill_blue = PatternFill(patternType='solid',fgColor=Color('000080'))

    for row in load_ws['B1':'G1']:
        for cell in row:
            cell.alignment = align_center
            cell.fill = fill_orange
            cell.border = border_thin


    for column in load_ws['A1':'A{}'.format(max_row)]:
        for cell in column:
            cell.alignment = align_center
            cell.fill = fill_orange
            cell.border = border_thin


    for row in load_ws['J1':'N1']:
        for cell in row:
            cell.alignment = align_center
            cell.fill = fill_blue
            cell.border = border_thin
            cell.font = font_white

    chart = BarChart()
    chart.title = "게시물 인사이트 요약"
    datas = Reference(load_ws, min_col=3,min_row=1,max_col=6,max_row=max_row)
    chart.add_data(datas, from_rows=False, titles_from_data=True)

    chart.width= 30
    chart.height = 10

    category = Reference(load_ws,min_col=7,min_row=2,max_col=7,max_row=max_row)
    chart.set_categories(category)

    load_ws.add_chart(chart, "J4")







wb = load_workbook('C:/Users/USER/Desktop/instagram.xlsx', data_only=True)


ws_names = wb.sheetnames

for i in range(len(ws_names)):
    ws = wb[ws_names[i]]
    xlsx_pre(ws)



ws = wb.create_sheet('{}'.format(day))
xlsx_decoration(wb,ws)

wb.save('C:/Users/USER/Desktop/instagram.xlsx')