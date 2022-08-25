from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
import pandas as pd
import time
import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side, Color, PatternFill
import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side, Color, PatternFill
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference , BarChart

export_time = datetime.datetime.now()
day = export_time.date()

path = "chromedriver.exe"
# url = "https://business.facebook.com/creatorstudio/login"

# url = "https://www.facebook.com/login/"


url = "https://www.facebook.com/%EC%8A%A4%ED%8A%9C%EB%94%94%EC%98%A4-%ED%99%80%EB%A6%B0-100815502019537"

id = "studiohollin@naver.com"
pw = "ghffls2020!"


sns_name=[]
like_lst = []
comment_lst=[]
reach_lst =[]
exposure_lst=[]
date_lst=[]


# 크롬 옵션 정의 (1이 허용, 2가 차단)
chrome_options = Options()
prefs = {"profile.default_content_setting_values.notifications": 2}
chrome_options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(path, options=chrome_options)
driver.get(url)


# 전체화면

driver.maximize_window()

# 로그인
driver.find_element("name", 'email').send_keys(id)
driver.find_element("name", 'pass').send_keys(pw)
driver.find_element("xpath",'/html/body/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/div/form/div[2]/div[3]/div/div').click()


# ActionChains 를 사용하기 위해서.
from selenium.webdriver import ActionChains


time.sleep(5)

i =1

try:

    while True:
        elem = driver.find_element("xpath",'/html/body/div[1]/div/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div[2]/div/div/div[4]/div[2]/div/div[2]/div[3]/div/div/div[{}]/div/div/div/div/div/div/div/div/div/div[2]/div/div[4]/div[2]'.format(i))
        actions = ActionChains(driver)
        actions.move_to_element(elem).perform()

        driver.find_element("xpath","/html/body/div[1]/div/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div[2]/div/div/div[4]/div[2]/div/div[2]/div[3]/div/div/div[{}]/div/div/div/div/div/div/div/div/div/div[2]/div/div[4]/div[2]".format(i)).click()


        time.sleep(3)

        reach = driver.find_element("xpath",'/html/body/div[1]/div[1]/div[1]/div/div[4]/div/div/div[1]/div/div[2]/div/div/div/div[3]/div[2]/div/div/div[1]/div[2]/div/div[1]/div/div/span').text
        like = driver.find_element("xpath",'/html/body/div[1]/div[1]/div[1]/div/div[4]/div/div/div[1]/div/div[2]/div/div/div/div[3]/div[2]/div/div/div[1]/div[3]/div/div[1]/div/div/span').text
        exposure = driver.find_element("xpath",'/html/body/div[1]/div[1]/div[1]/div/div[4]/div/div/div[1]/div/div[2]/div/div/div/div[3]/div[2]/div/div/div[1]/div[8]/div[3]/div[1]/div/div[1]/span').text
        date = driver.find_element("xpath",'/html/body/div[1]/div[1]/div[1]/div/div[4]/div/div/div[1]/div/div[2]/div/div/div/div[3]/div[1]/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div[2]/div/div[2]/span/span/span[2]/span/a/span').text



        if len(date)>5:
            date = date[0:6]

        sns_name.append('페이스북')
        reach_lst.append(reach)
        like_lst.append(like)
        exposure_lst.append(exposure)
        date_lst.append(date)

        print(reach,like,exposure,date)


        driver.find_element("xpath",'/html/body/div[1]/div[1]/div[1]/div/div[4]/div/div/div[1]/div/div[2]/div/div/div/div[2]/div').click()

        i+=1

except:


    facebook_record = pd.DataFrame({'활동영역':sns_name,
                                '노출': exposure_lst,
                                '도달': reach_lst,
                                '공감': like_lst,
                                '게시일':date_lst})



    facebook_record = facebook_record.astype({'노출': 'int', '도달': 'int', '공감': 'int'})


    print(facebook_record)

    driver.close()






    indicator = len(exposure_lst)

    def xlsx_pre(ws):

        load_ws = ws

        max_row = load_ws.max_row

        load_ws.cell(1, 10).value = "총 게시물"
        load_ws.cell(1, 11).value = "노출"
        load_ws.cell(1, 12).value = "도달"
        load_ws.cell(1, 13).value = "공감"
        # load_ws.cell(1, 14).value = "댓글"

        load_ws.cell(2, 10).value = '={}'.format(max_row - 1)
        load_ws.cell(2, 11).value = '=sum(c2:c{})'.format(max_row)
        load_ws.cell(2, 12).value = '=sum(d2:d{})'.format(max_row)
        load_ws.cell(2, 13).value = '=sum(e2:e{})'.format(max_row)
        # load_ws.cell(2, 14).value = '=sum(f2:f{})'.format(max_row)


    def xlsx_decoration(wb, ws):

        ws.cell(row=1, column=2).value = "활동영역"
        ws.cell(row=1, column=3).value = "노출"
        ws.cell(row=1, column=4).value = "도달"
        ws.cell(row=1, column=5).value = "공감"
        # ws.cell(row=1, column=6).value = "댓글"
        ws.cell(row=1, column=6).value = "게시일"

        for i in range(indicator):
            ws.cell(row=i + 2, column=1).value = i
            ws.cell(row=i + 2, column=2).value = "페이스북"
            ws.cell(row=i + 2, column=3).value = int(exposure_lst[i])
            ws.cell(row=i + 2, column=4).value = int(reach_lst[i])
            ws.cell(row=i + 2, column=5).value = int(like_lst[i])
            # ws.cell(row=i + 2, column=6).value = int(comment_lst[i])
            ws.cell(row=i + 2, column=6).value = date_lst[i]

        load_wb = wb
        load_ws = ws

        max_row = load_ws.max_row

        load_ws.cell(1, 10).value = "총 게시물"
        load_ws.cell(1, 11).value = "노출"
        load_ws.cell(1, 12).value = "도달"
        load_ws.cell(1, 13).value = "공감"
        # load_ws.cell(1, 14).value = "댓글"

        load_ws.cell(2, 10).value = '={}'.format(max_row - 1)
        load_ws.cell(2, 11).value = '=sum(c2:c{})'.format(max_row)
        load_ws.cell(2, 12).value = '=sum(d2:d{})'.format(max_row)
        load_ws.cell(2, 13).value = '=sum(e2:e{})'.format(max_row)
        # load_ws.cell(2, 14).value = '=sum(f2:f{})'.format(max_row)

        font_15 = Font(name='맑은 고딕', size=15, bold=True)
        font_white = Font(color='ffffff', bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        align_vcenter = Alignment(vertical='center')

        border_thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')

        )

        fill_orange = PatternFill(patternType='solid', fgColor=Color('ffe9b8'))
        fill_blue = PatternFill(patternType='solid', fgColor=Color('000080'))

        for row in load_ws['B1':'F1']:
            for cell in row:
                cell.alignment = align_center
                cell.fill = fill_orange
                cell.border = border_thin

        for column in load_ws['A1':'A{}'.format(max_row)]:
            for cell in column:
                cell.alignment = align_center
                cell.fill = fill_orange
                cell.border = border_thin

        for row in load_ws['J1':'M1']:
            for cell in row:
                cell.alignment = align_center
                cell.fill = fill_blue
                cell.border = border_thin
                cell.font = font_white

        chart = BarChart()
        chart.title = "게시물 인사이트 요약"
        datas = Reference(load_ws, min_col=3, min_row=1, max_col=5, max_row=max_row)
        chart.add_data(datas, from_rows=False, titles_from_data=True)

        chart.width = 30
        chart.height = 10

        category = Reference(load_ws, min_col=6, min_row=2, max_col=6, max_row=max_row)
        chart.set_categories(category)

        load_ws.add_chart(chart, "J4")


    wb = load_workbook('C:/Users/USER/Desktop/facebook.xlsx', data_only=True)

    ws_names = wb.sheetnames

    for i in range(len(ws_names)):
        ws = wb[ws_names[i]]
        xlsx_pre(ws)

    ws = wb.create_sheet('{}'.format(day))
    xlsx_decoration(wb, ws)

    wb.save('C:/Users/USER/Desktop/facebook.xlsx')



















